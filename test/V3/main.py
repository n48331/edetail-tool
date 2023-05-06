
# Tool: Janssen EDetail Generator
# Description: This tool helps in generating all required files by reading requirement from excel file .
# Author: Mohammed Nabeel P
# Version: 3.0.0(Development Tool)
import ctypes
import os
from file_handlers import createHtml, createConfig, renameSharedFiles, generateCssJs as gCJ, imageResize as IR, generateSharedHtml
import shutil
import openpyxl
import time
from pdf2image import convert_from_path
import zipfile
import colorama
import sys
stdoutOrigin=sys.stdout
sys.stdout = open("log.txt", "w")

start_time = time.time()


class Error(Exception):
    pass


class projectNameError(Error):
    pass


class imageFileError(Error):
    pass


class popupError(Error):
    pass


class pdfError(Error):
    pass
class requiredFilesError(Error):
    pass


def log(message, color=colorama.Fore.WHITE):
    if color == 1:
        color = colorama.Fore.RED
    elif color == 2:
        color = colorama.Fore.GREEN
    elif color == 3:
        color = colorama.Fore.BLUE

    # print(color + message + colorama.Style.RESET_ALL)
    print(message)


def Mbox(title, text, style):
    return ctypes.windll.user32.MessageBoxW(0, text, title, style)

try:
    book = openpyxl.load_workbook("./data_pdf.xlsx")
    data = book.active
    log('excel file accessed........', 2)
    project_name = data['A2'].value
    project_id = project_name.split('_')[0]

    log('---------------- Checking all required folders are present ----------------', 3)
    source_folder = "./boiler_plate/slide"
    image_folder = "./images"
    if not os.path.exists(image_folder):
        os.makedirs(image_folder)
        log("The folder 'images' created.")
    else:
        log(f"The folder {image_folder} already exists.")

    pdf_path = './pdf'
    pdf = ''
    pdf_filename ='not found'
    shared_src = "./boiler_plate/shared"
    shared_dest = f"./output/{project_name}/shared"
    config_dest = f"./output/{project_name}/shared/{project_id}_SharedResource/common"
    poppler_path = r'poppler-22.04.0\Library\bin'
    for filename in os.listdir(pdf_path):
        if filename.endswith(".pdf"):
            pdf = os.path.join(pdf_path, filename)
            pdf_filename = filename
            break
    if pdf == '':
        log("Cant continue without pdf. please place pdf in pdf folder in root path", 1)
        shutil.rmtree(image_folder)
        raise pdfError
    log(f'pdf folder with pdf : {pdf != ""} ({pdf_filename})',
        2 if pdf != "" else 1)
    log(f'source_folder : {os.path.exists(source_folder)} ({source_folder})',
        2 if os.path.exists(source_folder) else 1)
    log(f'image folder : {os.path.exists(image_folder)} ({image_folder})',
        2 if os.path.exists(image_folder) else 1)
    log(f'shared_src : {os.path.exists(shared_src)} ({shared_src})',
        2 if os.path.exists(shared_src) else 1)
    log(f'poppler : {os.path.exists(poppler_path)} ({poppler_path})',
        2 if os.path.exists(poppler_path) else 1)
    if pdf == '' or not os.path.exists(shared_src) or not os.path.exists(image_folder) or not os.path.exists(source_folder) or not os.path.exists(poppler_path):
        log("Cant continue without above file(s)", 1)
        raise requiredFilesError
    log('---------------- Checking Complete ----------------', 3)
    if project_name == None or project_name == '':
        raise projectNameError
    log(f'project id: {project_id}')
    log(f'project name: {project_name}')

    log('---------------- Searching and converting pdf ----------------', 3)


    if not os.path.exists(pdf):
        log('pdf not found. please upload pdf in pdf folder.', 1)
        raise pdfError
    pdf_images = convert_from_path(
        pdf, poppler_path=poppler_path, size=(2048, 1536))
    log(f"'{pdf_filename}' with {len(pdf_images)} pages fetched...........", 2)

    for i in range(len(pdf_images)):
        pdf_images[i].save(image_folder+'/Slide_' + str(i+1) + '.jpg', 'JPEG')
        log('Extracted Page' + str(i+1) + ' image')
    log('--------------- images saved in images folder -------------', 3)

    slide_names = []
    carousal_slides = []
    max_row = 0
    for row in data:
        if not all([cell.value is None for cell in row]):
            max_row += 1
    log(f'----------------{max_row-1} KMs-----------------', 3)
    for i in range(max_row-1):
        folder_name = f'{project_name}_S{str(i+1)}_{(data[f"B{i+2}"].value).replace(" ","_")}'

        slide_names.append(folder_name)
        popups = data[f'D{i+2}'].value
        buttons = data[f'E{i+2}'].value
        slides = str(data[f"C{i+2}"].value).split(',')
        slides = [f'Slide_{x}.jpg' for x in slides]
        if len(slides) > 1:
            carousal_slides.append(f"s{i+1}")
        if (popups != None):
            popups = str(popups).split(',')
            popups = [f'Slide_{x}.jpg' for x in popups]
        else:
            popups = ''
        if (buttons != None):
            buttons = buttons.split(',')
        else:
            buttons = ''

        log(
            f'--------------- KM {i+1}- {len(slides)} slide(s) | {len(popups)} popup(s) ---------------', 3)
        destination_folder = f"./output/{project_name}/{folder_name}"

        if not os.path.exists(destination_folder):
            shutil.copytree(source_folder, destination_folder)

        createHtml(project_id, folder_name,
                destination_folder, i+1, slides, len(popups), buttons)
        log('HTML files and required folders created.......', 2)
        gCJ(destination_folder, slides, popups, i+1)
        log('Generated CSS and JS..........', 2)
        # ========= Images ===========
        try:
            for slide in slides:
                img_name = f'{image_folder}/{slide}'
                IR(img_name, (2048, 1536), f'{destination_folder}/img/{slide}')
                log(f'{slide} image resized and saved.......')
            img_name = f'{image_folder}/{slides[0]}'
            IR(img_name, (1024, 768),
                f'{destination_folder}/{folder_name}-full.jpg')
            log('Full image resized and saved........')
            IR(img_name, (200, 150),
                f'{destination_folder}/{folder_name}-thumb.jpg')
            log('Thumb image resized and saved........')
        except:
            log('Slide, Full, Thumb images IR error.', 1)
            raise imageFileError
        # ========= popups ===========
        try:
            for i, popup in enumerate(popups):
                IR(f'{image_folder}/{popup}', (2048, 1536),
                    f'{destination_folder}/img/{popup}')
                log(f'Popup {i+1} image resized and saved.......')

        except:
            log('Popup image IR error', 1)
            raise popupError
    log('--------------- slides template complete -------------', 3)

    if not os.path.exists(shared_dest):
        shutil.copytree(shared_src, shared_dest)
    shared_image = f'Slide_{str(data["C2"].value)[0]}.jpg'
    sharedHTML = renameSharedFiles(shared_dest, project_id,
                                f'{image_folder}/{shared_image}')
    generateSharedHtml(project_id, sharedHTML)
    log('Shared resource folder created..........')
    createConfig(project_name, carousal_slides, slide_names, config_dest)
    log('config file added..............')

    # output km names
    output_xl = openpyxl.Workbook()
    output_xl_sheet = output_xl.active
    for i, slide in enumerate(slide_names):
        output_xl_sheet.cell(row=i+1, column=1).value = slide
    output_xl.save(f'./output/{project_name}/km_names.xlsx')
    log('KM names excel file generated', 2)
    shutil.rmtree(image_folder)
    log('image folder deleted...........', 1)
    with zipfile.ZipFile(f'./output/{project_id}.zip', "w", zipfile.ZIP_DEFLATED) as zip_obj:
        for foldername, subfolders, filenames in os.walk(f'./output/{project_name}'):
            for filename in filenames:
                file_path = os.path.join(foldername, filename)
                zip_obj.write(file_path)
    log('Project folder compressed...........', 2)
    shutil.rmtree(f'./output/{project_name}')
    log('Project output folder deleted...........', 1)
    os.remove(pdf)
    log('Source pdf deleted...........', 1)
    exec_time = "%s seconds." % round((time.time() - start_time),2)
    log('Total proccess time: '+ exec_time)
    log('Process Finished', 2)



except pdfError:
    print('ERROR : PDF - Please upload pdf')
    Mbox('ERROR',
         "PDF - Please upload pdf, please check process log for more details", 0x30)
    exit(0)
except requiredFilesError:
    print('ERROR : Files/Folders - Some required files are missing')
    Mbox('ERROR',
         "Files/Folders - Some required files are missing, please check process log for more details", 0x30)
    exit(0)
except imageFileError:
    print('ERROR : images - Check slide names in Excel or "images" folder is missing')
    Mbox('ERROR',
         "images - Check slide names in Excel or 'images' folder is missing, please check process log for more details", 0x30)
    exit(0)
    
except projectNameError:
    print('ERROR : Project name should not be empty in Excel')
    Mbox('ERROR',
         "Project name should not be empty in Excel, please check process log for more details", 0x30)
    exit(0)
except popupError:
    print('ERROR : Popup images are missing or wrong names in Excel ')
    Mbox('ERROR',
         "Popup images are missing or wrong names in Excel, please check process log for more details", 0x30)
    exit(0)
except FileNotFoundError:
    print(
        "ERROR : Make sure ['images','boiler_plate','data.xlsx'] are present in same folder or some files are missing in boiler_plate")
    Mbox('ERROR',
         "Make sure ['images','boiler_plate','data.xlsx'] are present in same folder or some files are missing in boiler_plate folder.check readme file for folder stucture, please check process log for more details", 0x20)
    exit(0)

except:
    Mbox(f'ERROR !',
         'Something went wrong....., please check process log for more details', 0x20)
    exit(0)

print('Process Finished')
colorama.deinit()
sys.stdout.close()
sys.stdout=stdoutOrigin
Mbox(f'Hurray Process Finished successfully ! in {exec_time}',
     'Check output folder for output, process log available in root path', 0x40)
